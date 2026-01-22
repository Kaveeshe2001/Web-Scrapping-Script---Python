from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Load workbook and select sheet
wb = load_workbook('pivot_table_output.xlsx')
sheet = wb['Report']

#Get data boundaries
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Create Bar Chart
barchart = BarChart()
barchart.style = 10

# Data & Categories
data = Reference( sheet, min_col=min_column + 1, max_col=max_column - 1, min_row=min_row, max_row=max_row - 1 )
categories = Reference( sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row - 1 )

# Add data & categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

barchart.title = "Sales by Product Line"

# Place chart in sheet
sheet.add_chart(barchart, "B12")

wb.save('barchart.xlsx')